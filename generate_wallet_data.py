#!/usr/bin/env python3
"""
Bitkey Address Generator
------------------------
Derive Native SegWit (P2WSH) multisig addresses from Bitkey-style
sortedmulti descriptors, optionally query balances / transactions via
a Bitcoin Core JSON-RPC node, and export results to Excel.
"""

from __future__ import annotations

import argparse
import hashlib
import logging
import os
import re
import sys
from collections import defaultdict
from typing import Any

import openpyxl
from bech32 import bech32_encode, convertbits
from bip32utils import BIP32Key
from bitcoinlib.services.authproxy import AuthServiceProxy, JSONRPCException
from dotenv import load_dotenv

logger = logging.getLogger("bitkey_address_generator")


# ---------------------------------------------------------------------------
# Descriptor & address derivation
# ---------------------------------------------------------------------------

def parse_descriptor(descriptor: str) -> tuple[int, list[dict[str, str]]]:
    """Parse a sortedmulti(...) descriptor into (m, [{fingerprint, xpub}, ...])."""
    match = re.search(r"sortedmulti\((\d+),(.+)\)", descriptor)
    if not match:
        raise ValueError(
            "Descriptor format is incorrect — expected sortedmulti(m,key1,key2,...)"
        )

    required_signatures = int(match.group(1))
    keys = match.group(2).split(",")
    parsed_keys: list[dict[str, str]] = []

    for key in keys:
        fingerprint_match = re.search(r"\[(.+?)\]", key)
        xpub_match = re.search(r"\](.+)/\d+/\*", key)
        if fingerprint_match and xpub_match:
            parsed_keys.append(
                {
                    "fingerprint": fingerprint_match.group(1),
                    "xpub": xpub_match.group(1),
                }
            )

    if not parsed_keys:
        raise ValueError("No valid keys found in descriptor")

    return required_signatures, parsed_keys


def derive_bech32_address(
    pubkeys: list[bytes],
    required_signatures: int,
    script_type: str = "wsh",
) -> str:
    """Build a mainnet Bech32 address (P2WSH multisig or P2WPKH)."""
    if script_type == "wsh":
        n = len(pubkeys)
        m = required_signatures
        # BIP67: sorted pubkeys inside the witness script
        witness_script = (
            bytes([0x50 + m])
            + b"".join(bytes([0x21]) + pk for pk in sorted(pubkeys))
            + bytes([0x50 + n, 0xAE])  # OP_n OP_CHECKMULTISIG
        )
        witness_program = hashlib.sha256(witness_script).digest()
        logger.debug("Witness script: %s", witness_script.hex())
        logger.debug("Witness program: %s", witness_program.hex())
    elif script_type == "wpkh":
        witness_program = hashlib.new(
            "ripemd160", hashlib.sha256(pubkeys[0]).digest()
        ).digest()
        logger.debug("P2WPKH witness program: %s", witness_program.hex())
    else:
        raise ValueError(f"Unsupported script type: {script_type}")

    program = convertbits(witness_program, 8, 5, True)
    if program is None:
        raise ValueError("Failed to convert witness program to 5-bit groups")

    # Witness version 0 → classic Bech32 (not Bech32m)
    address = bech32_encode("bc", [0] + program)
    if not address:
        raise ValueError("Bech32 encoding failed")
    return address


def derive_addresses(
    descriptor: str,
    is_change: bool = False,
    count: int = 20,
) -> list[str]:
    """
    Derive `count` addresses from a multisig descriptor.

    External (receive) uses chain index 0; internal (change) uses 1.
    Child index i is the address index within that chain.
    """
    path_index = 1 if is_change else 0
    required_signatures, parsed_keys = parse_descriptor(descriptor)
    addresses: list[str] = []

    for i in range(count):
        pubkeys = []
        for key_info in parsed_keys:
            key = BIP32Key.fromExtendedKey(key_info["xpub"])
            child_key = key.ChildKey(path_index).ChildKey(i)
            pubkeys.append(child_key.PublicKey())
        addresses.append(
            derive_bech32_address(pubkeys, required_signatures, script_type="wsh")
        )

    return addresses


# ---------------------------------------------------------------------------
# Bitcoin Core RPC helpers
# ---------------------------------------------------------------------------

def make_rpc(node_url: str, username: str, password: str, timeout: int = 60):
    rpc_url = f"http://{username}:{password}@{node_url}"
    return AuthServiceProxy(rpc_url, timeout=timeout)


def abort_scan(rpc_connection, attempts: int = 3) -> None:
    for _ in range(attempts):
        try:
            rpc_connection.scantxoutset("abort")
            logger.info("Aborted prior scantxoutset successfully.")
            return
        except JSONRPCException as e:
            if "no scan in progress" in str(e).lower():
                return
            logger.warning("Retrying scan abort: %s", e)


def fetch_balances_via_scan(
    rpc_connection,
    addresses: list[str],
) -> dict[str, float]:
    """Sum UTXO amounts per address via scantxoutset."""
    abort_scan(rpc_connection)
    scan_objects = [f"addr({address})" for address in addresses]
    scan_result = None

    for attempt in range(3):
        try:
            scan_result = rpc_connection.scantxoutset("start", scan_objects)
            break
        except Exception as e:
            logger.warning("scantxoutset attempt %s failed: %s", attempt + 1, e)
            if attempt == 2:
                raise

    if not scan_result or "unspents" not in scan_result:
        raise ValueError("Unexpected response from scantxoutset")

    balances = {address: 0.0 for address in addresses}
    for utxo in scan_result["unspents"]:
        # desc looks like "addr(<address>)#checksum" or "addr(<address>)"
        desc = utxo.get("desc", "")
        address = desc[5:].split(")")[0] if desc.startswith("addr(") else ""
        if address in balances:
            balances[address] += float(utxo.get("amount", 0))

    return balances


def setup_descriptor_wallet(
    rpc_connection,
    wallet_name: str,
    external_descriptor: str,
    internal_descriptor: str,
    address_range: tuple[int, int] = (0, 100),
):
    """Create (if needed) a watch-only descriptor wallet and import descriptors."""
    existing = rpc_connection.listwallets()
    if wallet_name not in existing:
        logger.info("Creating watch-only descriptor wallet: %s", wallet_name)
        rpc_connection.createwallet(
            wallet_name,
            disable_private_keys=True,
            descriptors=True,
        )

    wallet_rpc = rpc_connection.wallet(wallet_name)
    import_payload = [
        {
            "desc": external_descriptor,
            "timestamp": "now",
            "active": True,
            "label": "external",
            "range": list(address_range),
            "watchonly": True,
        },
        {
            "desc": internal_descriptor,
            "timestamp": "now",
            "active": True,
            "internal": True,
            "range": list(address_range),
            "watchonly": True,
        },
    ]
    wallet_rpc.importdescriptors(import_payload)
    logger.info("Descriptors imported into wallet '%s'.", wallet_name)
    return wallet_rpc


def fetch_wallet_transactions(
    rpc_connection,
    known_addresses: dict[str, str],
    count: int = 100,
) -> list[dict[str, Any]]:
    """Summarize recent wallet transactions, tagging known addresses."""
    try:
        transactions = rpc_connection.listtransactions("*", count)
    except Exception as e:
        logger.error("Error listing transactions: %s", e)
        return []

    grouped: dict[str, list] = defaultdict(list)
    for tx in transactions:
        grouped[tx["txid"]].append(tx)

    summarized: list[dict[str, Any]] = []
    for txid, _tx_group in grouped.items():
        try:
            full_tx = rpc_connection.gettransaction(txid)
        except Exception as e:
            logger.warning("Could not load tx %s: %s", txid, e)
            continue

        details = full_tx.get("details", [])
        confirmations = full_tx.get("confirmations", 0)
        time = full_tx.get("time")

        inputs, outputs = [], []
        input_roles, output_roles = [], []
        total_amount = 0.0

        for d in details:
            addr = d.get("address", "")
            role = known_addresses.get(addr, "Unknown")
            if d.get("category") == "send":
                inputs.append(addr)
                input_roles.append(role)
                total_amount += float(d.get("amount", 0))
            elif d.get("category") == "receive":
                outputs.append(addr)
                output_roles.append(role)
                total_amount += float(d.get("amount", 0))

        is_change = any("Internal" in r for r in output_roles)
        summarized.append(
            {
                "txid": txid,
                "amount": total_amount,
                "inputs": [f"{a} ({r})" for a, r in zip(inputs, input_roles)],
                "outputs": [f"{a} ({r})" for a, r in zip(outputs, output_roles)],
                "change_detected": is_change,
                "confirmations": confirmations,
                "time": time,
            }
        )

    return summarized


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(
    external_addresses: list[str],
    internal_addresses: list[str],
    external_balances: dict[str, float],
    internal_balances: dict[str, float],
    tx_summaries: list[dict[str, Any]] | None,
    output_file: str,
) -> None:
    workbook = openpyxl.Workbook()

    sheet1 = workbook.active
    sheet1.title = "External Addresses"
    sheet1.append(["Index", "Address", "Balance (BTC)"])
    for idx, addr in enumerate(external_addresses, start=1):
        sheet1.append([idx, addr, external_balances.get(addr, 0)])

    sheet2 = workbook.create_sheet(title="Internal Addresses")
    sheet2.append(["Index", "Address", "Balance (BTC)"])
    for idx, addr in enumerate(internal_addresses, start=1):
        sheet2.append([idx, addr, internal_balances.get(addr, 0)])

    if tx_summaries is not None:
        tx_sheet = workbook.create_sheet(title="Transaction Summary")
        tx_sheet.append(
            [
                "TXID",
                "Amount",
                "Inputs",
                "Outputs",
                "Change Detected",
                "Confirmations",
                "Time",
            ]
        )
        for tx in tx_summaries:
            tx_sheet.append(
                [
                    tx["txid"],
                    tx["amount"],
                    ", ".join(tx["inputs"]),
                    ", ".join(tx["outputs"]),
                    "Yes" if tx["change_detected"] else "No",
                    tx["confirmations"],
                    tx["time"],
                ]
            )

    workbook.save(output_file)
    logger.info("Wrote Excel report → %s", output_file)


# ---------------------------------------------------------------------------
# Input parsing
# ---------------------------------------------------------------------------

def load_descriptors(input_file: str) -> tuple[str, str]:
    """
    Read external + internal descriptors from a text file.

    Expected format (two lines):
        External: wsh(sortedmulti(...))#checksum
        Internal: wsh(sortedmulti(...))#checksum
    """
    with open(input_file, "r", encoding="utf-8") as file:
        lines = [ln.strip() for ln in file.readlines() if ln.strip()]

    if len(lines) < 2:
        raise ValueError(
            "Input file needs at least two non-empty lines "
            "(External: ... and Internal: ...)"
        )

    external_line = next((ln for ln in lines if ln.lower().startswith("external:")), None)
    internal_line = next((ln for ln in lines if ln.lower().startswith("internal:")), None)

    if not external_line or not internal_line:
        raise ValueError(
            "Could not find 'External:' and 'Internal:' lines in the descriptor file"
        )

    external_descriptor = external_line.split(":", 1)[1].strip()
    internal_descriptor = internal_line.split(":", 1)[1].strip()
    return external_descriptor, internal_descriptor


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run(
    input_file: str,
    output_file: str,
    node_url: str,
    username: str,
    password: str,
    *,
    address_count: int = 20,
    dry_run: bool = False,
    skip_transactions: bool = False,
    wallet_name: str = "bitkey_watchonly",
) -> None:
    external_descriptor, internal_descriptor = load_descriptors(input_file)
    logger.info("Loaded descriptors from %s", input_file)

    external_addresses = derive_addresses(
        external_descriptor, is_change=False, count=address_count
    )
    internal_addresses = derive_addresses(
        internal_descriptor, is_change=True, count=address_count
    )

    logger.info("Derived %s external addresses", len(external_addresses))
    logger.info("Derived %s internal (change) addresses", len(internal_addresses))
    for i, addr in enumerate(external_addresses):
        logger.debug("External[%s] %s", i, addr)
    for i, addr in enumerate(internal_addresses):
        logger.debug("Internal[%s] %s", i, addr)

    if dry_run:
        print("\n=== Dry run — address derivation only ===\n")
        print("External (receive):")
        for i, addr in enumerate(external_addresses):
            print(f"  {i:>3}  {addr}")
        print("\nInternal (change):")
        for i, addr in enumerate(internal_addresses):
            print(f"  {i:>3}  {addr}")
        print("\nNo balances fetched, no Excel written.")
        return

    rpc = make_rpc(node_url, username, password)
    external_balances = fetch_balances_via_scan(rpc, external_addresses)
    internal_balances = fetch_balances_via_scan(rpc, internal_addresses)

    known_addresses = {addr: "External" for addr in external_addresses}
    known_addresses.update(
        {addr: "Internal (Change)" for addr in internal_addresses}
    )

    tx_summaries: list[dict[str, Any]] | None = None
    if not skip_transactions:
        try:
            wallet_rpc = setup_descriptor_wallet(
                rpc,
                wallet_name,
                external_descriptor,
                internal_descriptor,
                address_range=(0, max(address_count, 100)),
            )
            tx_summaries = fetch_wallet_transactions(wallet_rpc, known_addresses)
            logger.info("Collected %s transaction summaries", len(tx_summaries))
        except Exception as e:
            logger.warning(
                "Transaction summary skipped (wallet/RPC issue): %s", e
            )
            tx_summaries = []

    export_to_excel(
        external_addresses,
        internal_addresses,
        external_balances,
        internal_balances,
        tx_summaries,
        output_file,
    )

    ext_total = sum(external_balances.values())
    int_total = sum(internal_balances.values())
    print(f"\nExternal balance total: {ext_total:.8f} BTC")
    print(f"Internal balance total: {int_total:.8f} BTC")
    print(f"Combined total:         {ext_total + int_total:.8f} BTC")
    print(f"Report written to:      {output_file}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Derive Bitkey multisig addresses from wallet descriptors, "
            "optionally scan balances via Bitcoin Core RPC, and export Excel."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "-i",
        "--input",
        default=os.getenv("INPUT_FILE", "wallet_descriptors.txt"),
        help="Path to descriptor file (External: / Internal: lines)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=os.getenv("OUTPUT_FILE", "wallet_data.xlsx"),
        help="Excel output path",
    )
    parser.add_argument(
        "--node-url",
        default=os.getenv("NODE_URL", "127.0.0.1:8332"),
        help="Bitcoin Core RPC host:port",
    )
    parser.add_argument(
        "--rpc-user",
        default=os.getenv("RPC_USER", ""),
        help="Bitcoin Core RPC username (or set RPC_USER in .env)",
    )
    parser.add_argument(
        "--rpc-password",
        default=os.getenv("RPC_PASSWORD", ""),
        help="Bitcoin Core RPC password (or set RPC_PASSWORD in .env)",
    )
    parser.add_argument(
        "-n",
        "--count",
        type=int,
        default=int(os.getenv("ADDRESS_COUNT", "20")),
        help="How many addresses to derive per chain (external / change)",
    )
    parser.add_argument(
        "--wallet-name",
        default=os.getenv("WALLET_NAME", "bitkey_watchonly"),
        help="Name of the watch-only descriptor wallet created on the node",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only derive and print addresses — no RPC, no Excel",
    )
    parser.add_argument(
        "--skip-transactions",
        action="store_true",
        help="Skip descriptor-wallet import and transaction summary sheet",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Enable debug logging",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    load_dotenv()
    # Build parser after load_dotenv so env defaults are available
    parser = build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    if not args.dry_run and (not args.rpc_user or not args.rpc_password):
        logger.error(
            "RPC credentials required unless --dry-run is set. "
            "Pass --rpc-user/--rpc-password or set RPC_USER/RPC_PASSWORD in .env"
        )
        return 1

    if not os.path.isfile(args.input):
        logger.error("Descriptor file not found: %s", args.input)
        return 1

    try:
        run(
            input_file=args.input,
            output_file=args.output,
            node_url=args.node_url,
            username=args.rpc_user,
            password=args.rpc_password,
            address_count=args.count,
            dry_run=args.dry_run,
            skip_transactions=args.skip_transactions,
            wallet_name=args.wallet_name,
        )
    except Exception as e:
        logger.error("%s", e)
        if args.verbose:
            logger.exception("Full traceback:")
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
